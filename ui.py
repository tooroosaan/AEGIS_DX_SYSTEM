# pyinstaller --onefile UI/ui.py
# pyinstaller --onefile --noconsole --icon=C:\Users\ZHANG\PycharmProjects\AEGIS_DX_SYSTEM\UI\favicon.ico UI/ui.py
import sys
import os
import openpyxl
import unicodedata
import pandas as pd
from pandas import DataFrame, concat
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from PyQt5.QtWidgets import QMessageBox, QInputDialog, QApplication, QMainWindow, QVBoxLayout, QWidget, QPushButton, \
    QLabel, QFileDialog, QComboBox
from PyQt5.QtCore import Qt

list_desired = ['LLP', '北九州レストラン', '宮崎', '仙南', '広島事業所', '津山', '伊勢広域', '大村', '栃木', '愛西',
                '苫小牧', '県央みずほ',
                '美唄', '豊岡', '県央みずほ(売店)', '野洲川', '北九州', '小林', '加西市斎場', '半田市斎場', '松阪市斎場',
                '紫雲苑', '千葉市',
                '西都', '琵琶湖', 'みきやま', '富士宮', '北・鶴見', '吹田', '泉南阪南', '館林', '山形', '近江八幡', '豊中',
                '高島市', '大津',
                '三田', '津島', '福山', '天空館', '行田', '広呉', 'イージス', 'NSK', 'NDK', '皆野寄居', '名高速', '伊豆',
                '道路管理JV',
                '新見沼', 'ロテックス', 'セントラル']

# ============================================样式============================================================================
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# 设置背景颜色为橙色
orange_fill = PatternFill(start_color='F7CAAC', end_color='F7CAAC', fill_type='solid')
# 设置背景颜色为绿色
green_fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
# 设置背景颜色为灰色
gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

center_alignment = Alignment(horizontal='center', vertical='center')
# 设置字体
bold_font = Font(name='等线', size=10, bold=True)
font = Font(name='等线', size=10)


# ============================================================================================================================
class FileDropApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle('DXシステム')
        self.setGeometry(200, 200, 400, 400)
        self.central_widget = FileDropWidget()
        self.setCentralWidget(self.central_widget)


def process_text(value):
    # 使用unicodedata.normalize进行NFKC正规化
    if isinstance(value, str):
        return unicodedata.normalize('NFKC', value)
    else:
        return value


def adjust_worksheet_width(ws, plus=1.0, multiply=1.0):
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length * multiply) + plus  # 调整宽度的因子
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    return ws


def show_error_popup(error_message):
    error_popup = QMessageBox()
    error_popup.setIcon(QMessageBox.Critical)
    error_popup.setText("エラー")
    error_popup.setInformativeText(error_message)
    error_popup.setWindowTitle("エラー")
    error_popup.exec_()
    return error_message


# 创建一个自定义排序函数
def custom_sort_key(item):
    for word in list_desired:
        if word in item:
            # 返回list2中第一个匹配单词的位置，如果没有匹配单词，则返回 float('inf')
            return list_desired.index(word)
    return float('inf')


class FileDropWidget(QWidget):
    def __init__(self):
        super().__init__()
        self.process_button = None
        self.process_combo = None
        self.file_path = None
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        # 创建一个 QLabel 作为文件上传区域
        self.upload_area = QLabel('ここにファイルを置く\n\nもしくは\n\nここをクリックしてファイルを読み込む')
        self.upload_area.setAlignment(Qt.AlignCenter)
        # 使用 macOS 风格的样式
        self.upload_area.setStyleSheet("""
                    QLabel {
       
                        font:bold;
                        font-family: "Meiryo", monospace;
                        font-size: 30px;
                        border: 2px dashed #999;
                        padding: 20px;
                        background-color: #f4f4f4;
                        border-radius: 8px;
                        color: #333;
                    }
                """)
        layout.addWidget(self.upload_area)

        process_options = [
            "資金繰り表作成 (CSV to XLSX)",
            "資金繰り表追加 (XLSX＋CSV to XLSX)",
            "予実分析 (TXT to XLSX)",
            "出納帳作成 (CSV to XLSX)"
        ]
        self.process_combo = QComboBox()
        self.process_combo.addItems(process_options)

        # 设置 QComboBox 的样式为 macOS 风格
        self.process_combo.setStyleSheet("""
                  QComboBox {
                      font-family: "Courier New", monospace;
                      padding: 5px;
                      border: 1px solid #a8a8a8;
                      border-radius: 5px;
                      background-color: white;
                      selection-background-color: #a8a8a8; /* 选中项的背景色 */
                  }
                  QComboBox::drop-down {
                      subcontrol-origin: padding;
                      subcontrol-position: top right;
                      width: 20px;
                      border-left-width: 1px;
                      border-left-color: #a8a8a8;
                      border-left-style: solid;
                  }

                  QComboBox::down-arrow {
                      image: url(down_arrow.png); /* 自定义下拉箭头图标 */
                  }
              """)
        layout.addWidget(self.process_combo)

        self.process_button = QPushButton('処理する')
        self.process_button.setStyleSheet("""
                          QPushButton {
                              font-family: "Courier New", monospace;
                              padding: 5px;
                              border: 1px solid #a8a8a8;
                              border-radius: 5px;
                              background-color: white;
                              selection-background-color: #a8a8a8; /* 选中项的背景色 */
                          }
                          QPushButton:hover {
                              background-color: #e0e0e0; /* 悬停时的背景色 */
                          }
                          QPushButton:pressed {
                          background-color: #c0c0c0; /* 按下时的背景色 */
                          }
                          """)
        layout.addWidget(self.process_button)

        self.setLayout(layout)

        # 设置 QLabel 支持拖拽
        self.setAcceptDrops(True)
        self.process_button.clicked.connect(self.process_file)

        #     # 连接点击和拖拽事件的槽函数
        self.upload_area.mousePressEvent = self.browse_file
        self.upload_area.dragEnterEvent = self.dragEnterEvent
        self.upload_area.dropEvent = self.dropEvent

    def browse_file(self, event):
        # 点击上传区域触发浏览文件夹事件
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseCustomDirectoryIcons  # 显式设置为 True，确保使用原生对话框
        file_path, _ = QFileDialog.getOpenFileName(self, "ファイルを選択", "",
                                                   "ALL (*);;CSV (*.csv);;TXT (*.txt)", options=options)

        if file_path:
            self.upload_area.setText(f'選択したファイル：\n{file_path}')
            self.file_path = file_path
            self.process_button.setEnabled(True)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.accept()
        else:
            event.ignore()

    def dropEvent(self, event):
        for url in event.mimeData().urls():
            file_path = url.toLocalFile()
            self.upload_area.setText(f'置かれたファイル：\n{file_path}')

            self.file_path = file_path
            self.process_button.setEnabled(True)

    def process_file(self):
        selected_option = self.process_combo.currentText()

        if selected_option == "資金繰り表作成 (CSV to XLSX)":
            self.process_cashbook()
        elif selected_option == "資金繰り表追加 (XLSX＋CSV to XLSX)":
            self.process_cashbook_append()
        elif selected_option == "予実分析 (TXT to XLSX)":
            self.process_budget_control()
        elif selected_option == "出納帳作成 (CSV to XLSX)":
            self.process_accounting()

    # 在下载按钮的槽函数中添加默认文件名参数
    def download_excel_file(self, workbook, default_file_name=''):
        # 打开文件对话框以选择保存路径，并设置默认文件名
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseCustomDirectoryIcons  # 显式设置为 True，确保使用原生对话框
        save_path, _ = QFileDialog.getSaveFileName(self, "XLSXファイルをセーブパスを選んでください", default_file_name,
                                                   "XLSX (*.xlsx);;All Files (*)", options=options)
        if save_path:
            # 使用 openpyxl 保存工作簿
            workbook.save(save_path)
            self.upload_area.setText(f'XLSXファイルをセーブしました，セーブパス：\n{save_path}')

    def pre_process_data(self, df):
        df_copy = df.copy()

        df_copy = df_copy[df_copy['勘定科目'].str.contains('　')]
        df_copy['勘定科目'] = df_copy['勘定科目'].apply(process_text)

        df_copy["借方"] = df_copy["取引内容"]
        df_copy["貸方"] = df_copy["取引内容"]
        df_copy = df_copy[
            ["取引日", "借方金額", "借方", "貸方金額", "貸方", "残高", "勘定科目", "相手勘定科目", "相手部門"]]
        df_copy.loc[df_copy["借方金額"].isna(), "借方"] = ''
        df_copy.loc[df_copy["貸方金額"].isna(), "貸方"] = ''
        column_mapping = {'勘定科目': '口座名', '取引日': '日付'}
        df_copy.rename(columns=column_mapping, inplace=True)
        # 将日期列转换为日期时间类型
        df_copy['日付'] = pd.to_datetime(df_copy['日付'])
        return df_copy

    def beautify_subtable_worksheet(self, ws):
        """
        美化分表
        :param ws:
        :return:
        """
        # 获取表头行（第一行）和列名
        header_row = ws[1]
        column_names = [cell.value for cell in header_row]

        orange_column = column_names.index("借方金額") + 1  # 通过索引找到“借方”所在的列
        green_column = column_names.index("貸方金額") + 1  # 通过索引找到“贷方”所在的列
        date_column = column_names.index("日付") + 1  # 通过索引找到“日期”所在的列
        predict_column = column_names.index("予測") + 1  # 通过索引找到“日期”所在的列

        # 设置“借方”列下面的所有单元格为橙色
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=orange_column, max_col=orange_column):
            for cell in row:
                cell.fill = orange_fill

        # 设置“贷方”列下面的所有单元格为绿色
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=green_column, max_col=green_column):
            for cell in row:
                cell.fill = green_fill

        # 设置“预测”列下面的所有单元格为居中
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=predict_column, max_col=predict_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        number_format = '#,##0'
        # 遍历所有单元格，设置数值格式
        for row in ws.iter_rows():
            for cell in row:
                cell.number_format = number_format
                cell.border = border
                cell.font = font

        # 遍历所有单元格，设置日期格式
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=date_column, max_col=date_column):
            for cell in row:
                # 设置单元格的日期格式
                cell.number_format = 'yyyy-mm-dd'
                cell.alignment = Alignment(horizontal='left')
                cell.border = border
                cell.font = font

        # 循环遍历第一行的单元格，应用字体样式
        for cell in ws[1]:
            cell.fill = gray_fill
            cell.font = bold_font
        ws.freeze_panes = ws["B2"]
        return ws

    def beautify_sumtable_worksheet(self, ws):
        """
        美化总表
        :rtype: object
        :param ws:
        :return:
        """

        # 获取表头行（第一行）和列名
        header_row = ws[1]
        column_names = [cell.value for cell in header_row]

        date_column = column_names.index("日付") + 1  # 通过索引找到“日期”所在的列
        green_column = column_names.index("合計") + 1  # 通过索引找到“合计”所在的列

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=green_column, max_col=green_column):
            for cell in row:
                cell.fill = orange_fill

        number_format = '#,##0'
        # 遍历所有单元格，设置数值格式
        for row in ws.iter_rows():
            for cell in row:
                cell.number_format = number_format
                cell.border = border
                cell.font = font

        # 遍历所有单元格，设置日期格式
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=date_column, max_col=date_column):
            for cell in row:
                # 设置单元格的日期格式
                cell.number_format = 'yyyy-mm-dd'
                cell.alignment = Alignment(horizontal='left')
                cell.border = border
                cell.font = font

        # 遍历所有单元格，将负数的值改成红色
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    if cell.value < 0:
                        column_letter = get_column_letter(cell.column)
                        range_coordinate = f'{column_letter}{cell.row}'
                        ws[range_coordinate].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                                                fill_type="solid")  # 设置背景色为浅红色
                        cell.font = Font(color="FF0000", name='等线', size=10)  # 设置字体颜色为红色

        # 循环遍历第一行的单元格，应用字体样式
        for cell in ws[1]:
            cell.fill = gray_fill
            cell.font = bold_font
        ws.freeze_panes = ws["B2"]
        return ws

    def separate_df_by_account(self, df):
        """
        将df按照口座名分开，并形成account_dict，并传回account_name_list
        :param df:
        :return:
        """
        grouped = df.groupby('口座名')
        # 将每个分组保存为字典，其中键是分组的名称，值是相应的 DataFrame
        account_dict = {group: data for group, data in grouped}
        account_name_list = list(account_dict.keys())
        # 使用sorted函数，按照最后一个单词的顺序排序列表
        account_name_list = sorted(account_name_list, key=custom_sort_key)
        return account_dict, account_name_list

    def process_sub_table(self, account_dict, account_name_list):
        """
        处理子表，按照有前期繰越和没有前期繰越的表来处理，调整资金表顺序和重新计算残高。后面加了处理
        :param account_dict:
        :param account_name_list:
        :return:
        """
        # 通过访问sub_table字典中的不同分组来获取每个子表
        for i in range(len(account_name_list)):
            account_name = account_name_list[i]
            sub_table = account_dict[account_name]

            # 如果有前期繰越，则走1，反之走2号
            if sub_table.loc[sub_table["相手勘定科目"] == "前期繰越"].empty:
                sub_table = sub_table.sort_values(by=['日付', '借方金額', '貸方金額'], ascending=[True, False, False])
                sub_table.reset_index(drop=True, inplace=True)
                sub_table.loc[0, "残高"] = np.nan_to_num(sub_table.loc[0, "借方金額"])

                # 计算第二行到最后一行的"残高"
                for i in range(1, len(sub_table)):
                    sub_table.loc[i, '残高'] = np.nan_to_num(sub_table.loc[i - 1, '残高']) + np.nan_to_num(
                        sub_table.loc[i, '借方金額']) - np.nan_to_num(sub_table.loc[i, '貸方金額'])

                account_dict[account_name] = sub_table

            else:
                sub_table = sub_table.sort_values(by=['日付', '借方金額', '貸方金額'], ascending=[True, False, False])
                # 将满足条件的行移动到开头
                sub_table = concat([sub_table[sub_table['相手勘定科目'] == "前期繰越"],
                                    sub_table[sub_table['相手勘定科目'] != "前期繰越"]])
                sub_table.reset_index(drop=True, inplace=True)

                # 计算第二行到最后一行的"残高"
                for i in range(1, len(sub_table)):
                    sub_table.loc[i, '残高'] = np.nan_to_num(sub_table.loc[i - 1, '残高']) + np.nan_to_num(
                        sub_table.loc[i, '借方金額']) - np.nan_to_num(sub_table.loc[i, '貸方金額'])

                account_dict[account_name] = sub_table
        return account_dict

    def predict_sub_table(self, account_dict, account_name_list, end_date):
        end_date = pd.to_datetime(end_date)
        # 计算三个月前的日期后一天
        start_date = end_date + pd.DateOffset(days=1)
        start_date = start_date - pd.DateOffset(months=6)

        # 遍历
        for i in range(len(account_name_list)):
            account_name = account_name_list[i]
            # 得到子表
            sub_table = account_dict[account_name]
            sub_table_copy = sub_table.copy()

            forecasted_data = sub_table_copy.loc[
                (sub_table_copy['日付'] >= start_date) & (sub_table_copy['日付'] <= end_date)]

            if forecasted_data.empty:
                pass
            else:
                forecasted_data_copy = forecasted_data.copy()
                forecasted_data_copy.loc[:, '日付'] = forecasted_data_copy['日付'] + pd.DateOffset(months=6)
                forecasted_data_copy.loc[:, '予測'] = "✓"
                sub_table = pd.concat([sub_table, forecasted_data_copy], ignore_index=True, sort=False)

            # 计算第二行到最后一行的"残高"
            for i in range(1, len(sub_table)):
                sub_table.loc[i, '残高'] = np.nan_to_num(sub_table.loc[i - 1, '残高']) + np.nan_to_num(
                    sub_table.loc[i, '借方金額']) - np.nan_to_num(sub_table.loc[i, '貸方金額'])
            account_dict[account_name] = sub_table
        return account_dict

    def write_sub_table(self, account_dict, account_name_list, wb):
        """
        根据account_dict，将每个账号分表写入wb
        """
        # 通过访问sub_table字典中的不同分组来获取每个子表
        for i in range(len(account_name_list)):
            account_name = account_name_list[i]
            sub_table = account_dict[account_name]
            ws = wb.create_sheet(account_name)
            ws = self.write_df_into_worksheet(sub_table, ws)
            ws = self.write_predict_equation_into_worksheet(ws)
            ws = self.beautify_subtable_worksheet(ws)
            adjust_worksheet_width(ws, 8, 1.2)
        return wb

    def write_sum_table(self, account_name_list, wb, df):
        """
        根据account_dict，将每个账号分表写入wb
        """
        # 通过访问sub_table字典中的不同分组来获取每个子表
        ws = wb.create_sheet("合算")
        sum_table = self.create_sum_table(df)
        for i in range(len(account_name_list)):
            account_name = account_name_list[i]
            sum_table[account_name] = sum_table.apply(lambda
                                                          row: '=IFERROR(LOOKUP(2, 1/(((\'{sheet}\'!A:A) <= {date}) * (\'{sheet}\'!A:A) >= 40000), \'{sheet}\'!F:F),0)'.format(
                sheet=account_name, date="A" + str(row.name + 2)), axis=1)
        sum_table["合計"] = sum_table.apply(lambda row: '=SUM(B{date}:BJ{date})'.format(date=str(row.name + 2)), axis=1)
        self.write_df_into_worksheet(sum_table, ws)
        self.beautify_sumtable_worksheet(ws)
        adjust_worksheet_width(ws, 20, 0)
        return wb

    def create_sum_table(self, df):
        """
        该函数用于创建总表，并计算所有的10号与月末日期。
        :param df:
        :return:sum_table
        """
        # 找到最早月初日期
        earliest_month_start = df['日付'].min().date()
        latest_month_start = df['日付'].max().date()
        date_range = pd.date_range(start=earliest_month_start, end=latest_month_start, freq='D')
        # 提取每个月的十号和月末日期
        tenth_dates = date_range.to_period('M').to_timestamp() + pd.offsets.MonthBegin(0) + pd.Timedelta(days=9)
        end_of_month_dates = date_range.to_period('M').to_timestamp() + pd.offsets.MonthEnd(0)
        # 合并两个日期列表
        all_dates = sorted(list(set(tenth_dates).union(end_of_month_dates)))
        sum_table = pd.DataFrame({'日付': all_dates})
        # 打印 DataFrame
        return sum_table

    # def calculation_10th_and_month_end_balances(self, account_dict, account_name_list, sum_table):
    #     """
    #     计算每个账户的10号和月末的余额，并加入到总表sum_table中
    #     :param sum_table:
    #     :param account_dict:
    #     :param account_name_list:
    #     :return:sum_table
    #     """
    #     for i in range(len(account_name_list)):
    #         account_name = account_name_list[i]
    #         sub_table = account_dict[account_name]
    #         sum_table[account_name] = 0
    #         # 遍历每个日期，得到这个日期应有的余额
    #         for index, row in sum_table.iterrows():
    #             date_value = row['日付']
    #             last_balance_row = sub_table.loc[sub_table["日付"] <= date_value].tail(1)
    #             if not last_balance_row.empty:
    #                 sum_table.loc[index, account_name] = int(last_balance_row["残高"].iloc[0])
    #             else:
    #                 pass

        # sum_table['合計'] = sum_table.drop(columns=['日付']).sum(axis=1, skipna=True)
        # return sum_table

    def read_cashbook_to_account_dict(self, xlsx_file):
        wb = openpyxl.load_workbook(xlsx_file)
        account_name_list = wb.sheetnames
        account_name_list.remove("合算")
        account_dict = {}  # 创建一个空字典
        for i in range(len(account_name_list)):
            account_name = account_name_list[i]
            ws = wb[account_name]
            # 读取工作表的数据
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(row)
            # 将数据转换为pandas DataFrame，并设置第一行为表头
            cashbook = pd.DataFrame(data[1:], columns=data[0])

            if '予測' in cashbook.columns:
                cashbook = cashbook[cashbook['予測'] != "✓"]

            account_dict[account_name] = cashbook
        return account_dict, account_name_list

    def account_dict_to_df(self, account_dict, account_name_list):
        df = DataFrame()
        for i in range(len(account_name_list)):
            account_name = account_name_list[i]
            sub_table = account_dict[account_name]
            df = pd.concat([df, sub_table], ignore_index=True)
        return df

    def account_dict_append_new_data(self, account_dict, account_name_list, new_data_account_dict,
                                     new_data_account_name_list):

        # 判断两个name_list的不同，如果new_data_account_name_list多了新的物件，就把新的物件的sub_table接到account_dict里
        new_account_name_list = list(set(new_data_account_name_list) - set(account_name_list))
        if new_account_name_list:
            for i in new_account_name_list:
                account_dict[i] = new_data_account_dict[i]
            account_name_list = account_name_list + new_account_name_list
        else:
            pass

        # 现在要把新数据给贴到老数据里
        # 判断新的数据里面有没有前期躁越，如果有，就删掉，然后拼接，如果没有，就直接拼接。
        for i in new_data_account_name_list:
            sub_table = account_dict[i]
            new_sub_table = new_data_account_dict[i]
            # 判断日期是否有重合
            # 获取每个 DataFrame 中日期列的最小和最大日期
            sub_table_min_date, sub_table_max_date = sub_table['日付'].min(), sub_table['日付'].max()
            new_sub_table_min_date, new_sub_table_max_date = new_sub_table['日付'].min(), new_sub_table['日付'].max()

            # 检查是否存在重叠
            try:
                if sub_table_max_date < new_sub_table_min_date or sub_table_min_date > new_sub_table_max_date:
                    pass
                else:
                    raise ValueError("There is an overlap between the date of the new data and the old data.")
            except Exception as e:
                show_error_popup(f"An error occurred: {str(e)}")
                return e

            new_sub_table = new_sub_table[new_sub_table['相手勘定科目'] != '前日繰越']
            sub_table = pd.concat([sub_table, new_sub_table], ignore_index=True)

            account_dict[i] = sub_table

        return account_dict, account_name_list

    def write_df_into_worksheet(self, df, ws):
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=process_text(value))
        return ws

    def write_predict_equation_into_worksheet(self, ws):
        # 遍历每一行
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ws.max_column, max_col=ws.max_column):
            # 获取当前行的 '予測' 列的值
            cell_value = row[0].value

            # 如果 '予測' 列的值为 '✓'，则将残高设置成算式
            if cell_value == '✓':
                ws.cell(row=row[0].row, column=6,
                        value="=F" + str(row[0].row - 1) + "+B" + str(row[0].row) + "-D" + str(row[0].row))
        return ws

    def process_cashbook(self):
        try:
            # 处理 csv文件到 Excel 文件的逻辑
            csv_file = self.file_path
            # ==========================================读取文件=====================================================
            """
            PART 1
            """
            df = pd.read_csv(csv_file, delimiter=",", skiprows=1, encoding='cp932')
            wb = Workbook()
            # 预处理表格
            df = self.pre_process_data(df)

            # 找到最早的日期和最晚的日期
            earliest_date = df['日付'].min().date()
            latest_date = df['日付'].max().date()

            # # 处理分表
            account_dict, account_name_list = self.separate_df_by_account(df)
            account_dict = self.process_sub_table(account_dict, account_name_list)
            account_dict = self.predict_sub_table(account_dict, account_name_list, latest_date)
            self.write_sub_table(account_dict, account_name_list, wb)

            # 处理总表
            df = self.account_dict_to_df(account_dict, account_name_list)
            self.write_sum_table(account_name_list, wb, df)

            wb.remove(wb['Sheet'])
            save_excel_file = "資金繰り表（" + str(earliest_date) + "~" + str(latest_date) + "）.xlsx"

            # 在处理完成后调用下载函数
            self.download_excel_file(workbook=wb, default_file_name=save_excel_file)

            wb.close()

        except Exception as e:
            show_error_popup(f"An error occurred: {str(e)}")

    def process_cashbook_append(self):
        try:
            # 处理 TXT 文件到 Excel 文件的逻辑
            xlsx_file = self.file_path
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseCustomDirectoryIcons  # 显式设置为 True，确保使用原生对话框
            csv_file_path, _ = QFileDialog.getOpenFileName(self, "追加したいデータファイル（CSV）を選択", "",
                                                           "ALL FILES (*);;CSV (*.csv);",
                                                           options=options)

            if csv_file_path:
                self.csv_file_path = csv_file_path

            new_df = pd.read_csv(self.csv_file_path, delimiter=",", skiprows=1, encoding='cp932')
            latest_date = pd.to_datetime(new_df['取引日']).max().date()
            wb = Workbook()

            # 把xlsx文件读取成account_dict
            account_dict, account_name_list = self.read_cashbook_to_account_dict(xlsx_file)

            # 把account_dict中的每个子表都处理好
            account_dict = self.process_sub_table(account_dict, account_name_list)

            # 把csv文件读取成new_account_dict
            new_df = self.pre_process_data(new_df)

            new_data_account_dict, new_data_account_name_list = self.separate_df_by_account(new_df)

            # 这里需要加一个合并分表的操作
            account_dict, account_name_list = self.account_dict_append_new_data(account_dict, account_name_list,
                                                                                new_data_account_dict,
                                                                                new_data_account_name_list)

            # 把account_dict中的每个子表都处理好
            account_dict = self.process_sub_table(account_dict, account_name_list)
            account_dict = self.predict_sub_table(account_dict, account_name_list, latest_date)
            self.write_sub_table(account_dict, account_name_list, wb)

            # 处理总表
            df = self.account_dict_to_df(account_dict, account_name_list)
            self.write_sum_table(account_name_list, wb, df)

            earliest_date = df['日付'].min().date()

            wb.remove(wb['Sheet'])
            save_excel_file = "資金繰り表（" + str(earliest_date) + "~" + str(latest_date) + "）.xlsx"

            # 在处理完成后调用下载函数
            self.download_excel_file(workbook=wb, default_file_name=save_excel_file)
            wb.close()

        except Exception as e:
            error_message = show_error_popup(f"An error occurred: {str(e)}")

    def process_budget_control(self):
        try:
            # 处理 TXT 文件到 Excel 文件的逻辑
            txt_file = self.file_path

            # 输入从哪个地方开始删除
            cancel_project = "三ヶ根山スカイライン"
            # 读取文件
            df = pd.read_csv(txt_file, delimiter=",", encoding='SHIFT-JIS', skiprows=5)

            # 预处理数据
            # 删除无用列
            df.drop(['[表題行]', '部門'], axis=1, inplace=True)

            # 删除未从事的火葬场
            column_names_list = df.columns.tolist()
            index = column_names_list.index(cancel_project)

            unperformed_list = column_names_list[index:]

            df.drop(columns=unperformed_list, inplace=True)
            df.drop(columns="本社", inplace=True)

            # 删除行号范围为 258 到 270 的行
            rows_to_delete = list(range(0, 2)) + list(range(6, 24)) + list(range(27, 29)) + list(range(30, 54)) + list(
                range(63, 81)) + \
                             list(range(87, 96)) + list(range(105, 257)) + list(range(258, 270)) + list(
                range(277, 303)) + list(range(306, 309)) + \
                             list(range(315, 377)) + list(range(381, 410)) + list(range(411, 414)) + list(
                range(417, 426)) + list(range(445, 453)) + \
                             list(range(495, 498)) + list(range(507, 522)) + list(range(540, 543)) + list(
                range(549, 573))

            df.drop(rows_to_delete, inplace=True)

            # 加入合计列
            numeric_columns = df.select_dtypes(include=['int64', 'float64'])
            df["合計"] = numeric_columns.sum(axis=1)

            # 重新设置index
            df.reset_index(inplace=True, drop=True)

            # 创建一个 Excel 工作簿对象
            wb = Workbook()

            ws = wb.active
            ws.title = txt_file.split("/")[-1].split(".txt")[0] + "_予実表"

            # 将 DataFrame 数据写入工作表
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)
            # 调整单元格宽度
            adjust_worksheet_width(ws, 13, 1)

            # 定义要合并的列
            column_to_merge = 'A'
            # 获取列的最大行数
            max_row = ws.max_row
            # 逐行检查并合并单元格
            start_row = 1
            for row in range(2, max_row + 3):
                current_value = ws[column_to_merge + str(row)].value
                previous_value = ws[column_to_merge + str(row - 1)].value if row > 1 else None

                if current_value != previous_value:
                    end_row = row - 1

                    if start_row != end_row:
                        merge_range = f'{column_to_merge}{start_row}:{column_to_merge}{end_row}'
                        ws.merge_cells(merge_range)

                        # 设置合并后的单元格样式
                        merged_cell = ws[column_to_merge + str(start_row)]
                        merged_cell.alignment = Alignment(horizontal='left', vertical='center')

                    start_row = row

            # 冻结前两列
            ws.freeze_panes = ws["C2"]

            # 遍历所有单元格，设置字体样式为 "等线"
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = Font(name='等线')

            # 设置数值格式为千分位加逗号
            number_format = '#,##0'
            # 遍历所有单元格，设置数值格式
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = number_format

            # 遍历所有单元格，将负数的值改成红色
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        if cell.value < 0:
                            cell.font = Font(color="FF0000", name='等线')  # 设置字体颜色为红色

                            if (ws.cell(row=cell.row, column=2).value == "差異") and (cell.value < -300000):
                                column_letter = get_column_letter(cell.column)
                                range_coordinate = f'{column_letter}{cell.row}'
                                ws[range_coordinate].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                                                        fill_type="solid")  # 设置背景色为浅红色

                        if cell.value > 300000:
                            if ws.cell(row=cell.row, column=2).value == "差異":
                                column_letter = get_column_letter(cell.column)
                                range_coordinate = f'{column_letter}{cell.row}'

                                ws[range_coordinate].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                                                                        fill_type="solid")  # 设置背景色为浅红色
                                ws[range_coordinate].font = Font(color="006100", name='等线')

            # ==========================================================================

            df_analysis = df.loc[df["勘定科目"] == "税引前当期純損益金額"].loc[df["予算/実績/差異"] == "差異"]
            df_analysis.drop(["勘定科目", "予算/実績/差異", "合計"], axis=1, inplace=True)
            df_analysis = df_analysis.transpose()

            # 重新设置index
            df_analysis.reset_index(inplace=True)
            df_analysis.columns = ["物件名", "損益予実差"]
            df_analysis = df_analysis[(df_analysis["損益予実差"] >= 600000) | (df_analysis["損益予実差"] <= -600000)]

            # 提取第一张表中売上高为0的物件
            pickup_list = df.loc[df["勘定科目"] == "売上高"].loc[df["予算/実績/差異"] == "予算"]
            pickup_list = df.columns[pickup_list.eq(0).all()]

            df_pickup = df_analysis[df_analysis['物件名'].isin(pickup_list)]
            df_analysis = df_analysis[~df_analysis['物件名'].isin(pickup_list)]

            # 重新设置index
            df_analysis.reset_index(drop=True, inplace=True)
            df_pickup.reset_index(drop=True, inplace=True)

            # 将物件分成火葬场和道路
            cut_name = "広島呉　坂"
            # 使用布尔索引来定位符合条件的行
            matching_rows = df_analysis[df_analysis["物件名"] == cut_name].index.tolist()[0]
            df_analysis_upper = df_analysis[:matching_rows]

            # 下部分的 DataFrame，包含分界线之后的行
            df_analysis_lower = df_analysis[matching_rows:]

            # ===================================================================
            """
            PART 2
            """
            # 这一部分是创建一个新的分析表，表里要包含所有物件的差，以及每个物件的具体情况

            # 创建新的工作表
            title2 = txt_file.split("/")[-1].split(".txt")[0] + " 予実分析"
            print(title2)
            # title2 = "予実分析"
            ws2 = wb.create_sheet(title2)  # 指定名称为 "Sheet2" 的工作表

            """
            左边的部分
            """

            # 将df_analysis写进ws里
            ws2.cell(row=1, column=1, value="物件名").font = Font(name='等线')
            ws2.cell(row=1, column=2, value="損益予実差").font = Font(name='等线')

            # 火葬场部分
            for row_idx, row_values in enumerate(df_analysis_upper.values, 2):
                for col_idx, col_value in enumerate(row_values, 1):
                    if isinstance(col_value, float) and (col_value < 0):
                        ws2.cell(row=row_idx, column=col_idx, value=col_value).font = Font(color="FF0000", name='等线')
                    else:
                        ws2.cell(row=row_idx, column=col_idx, value=col_value).font = Font(name='等线')

            # 计算列的合计值
            col_sum = df_analysis_upper["損益予実差"].sum()

            ws2.cell(row=len(df_analysis_upper) + 2, column=3, value=col_sum).font = Font(name='等线')

            # 加上划线
            ws2.cell(row=len(df_analysis_upper) + 2, column=1).border = Border(top=Side(border_style='thin'))
            ws2.cell(row=len(df_analysis_upper) + 2, column=2).border = Border(top=Side(border_style='thin'))
            ws2.cell(row=len(df_analysis_upper) + 2, column=3).border = Border(top=Side(border_style='thin'))

            # 道路部分
            for row_idx, row_values in enumerate(df_analysis_lower.values, 2):
                for col_idx, col_value in enumerate(row_values, 1):
                    if isinstance(col_value, float) and (col_value < 0):
                        ws2.cell(row=len(df_analysis_upper) + 3 + row_idx, column=col_idx, value=col_value).font = Font(
                            color="FF0000", name='等线')
                    else:
                        ws2.cell(row=len(df_analysis_upper) + 3 + row_idx, column=col_idx, value=col_value).font = Font(
                            name='等线')

            # 计算列的合计值
            col_sum = df_analysis_lower["損益予実差"].sum()
            ws2.cell(row=len(df_analysis_upper) + len(df_analysis_lower) + 5, column=3, value=col_sum).font = Font(
                name='等线')

            # 加上划线
            ws2.cell(row=len(df_analysis_upper) + len(df_analysis_lower) + 5, column=1).border = Border(
                top=Side(border_style='thin'))
            ws2.cell(row=len(df_analysis_upper) + len(df_analysis_lower) + 5, column=2).border = Border(
                top=Side(border_style='thin'))
            ws2.cell(row=len(df_analysis_upper) + len(df_analysis_lower) + 5, column=3).border = Border(
                top=Side(border_style='thin'))

            # 列出预算为0的项目
            ws2.cell(row=len(df_analysis_upper) + len(df_analysis_lower) + 8, column=1,
                     value="予算未作成物件").font = Font(
                name='等线')

            # 加上划线
            ws2.cell(row=len(df_analysis_upper) + len(df_analysis_lower) + 9, column=1).border = Border(
                top=Side(border_style='thin'))
            ws2.cell(row=len(df_analysis_upper) + len(df_analysis_lower) + 9, column=2).border = Border(
                top=Side(border_style='thin'))
            ws2.cell(row=len(df_analysis_upper) + len(df_analysis_lower) + 9, column=3).border = Border(
                top=Side(border_style='thin'))

            # 预算未完成项目
            for row_idx, row_values in enumerate(df_pickup.values, 0):
                for col_idx, col_value in enumerate(row_values, 1):
                    if isinstance(col_value, float) and (col_value < 0):
                        ws2.cell(row=len(df_analysis_upper) + len(df_analysis_lower) + 9 + row_idx, column=col_idx,
                                 value=col_value).font = Font(color="FF0000", name='等线')
                    else:
                        ws2.cell(row=len(df_analysis_upper) + len(df_analysis_lower) + 9 + row_idx, column=col_idx,
                                 value=col_value).font = Font(name='等线')
                    # 加上划线
            ws2.cell(row=len(df_analysis_upper) + len(df_analysis_lower) + len(df_pickup) + 9,
                     column=1).border = Border(
                top=Side(border_style='thin'))
            ws2.cell(row=len(df_analysis_upper) + len(df_analysis_lower) + len(df_pickup) + 9,
                     column=2).border = Border(
                top=Side(border_style='thin'))
            ws2.cell(row=len(df_analysis_upper) + len(df_analysis_lower) + len(df_pickup) + 9,
                     column=3).border = Border(
                top=Side(border_style='thin'))
            # 计算列的合计值
            col_sum = df_pickup["損益予実差"].sum()
            ws2.cell(row=len(df_analysis_upper) + len(df_analysis_lower) + len(df_pickup) + 9, column=3,
                     value=col_sum).font = Font(name='等线')
            # =========================================

            """
            右边的部分
            """
            # 列出所有需要分析的物件，形成list
            analysis_list = df_analysis['物件名'].tolist()

            # 计算每个物件的问题
            df_analysis_sub = df.loc[df["予算/実績/差異"] == "差異"].copy()
            df_analysis_sub.drop(["予算/実績/差異", "合計"], axis=1, inplace=True)
            # 重新设置index
            df_analysis_sub.reset_index(drop=True, inplace=True)

            # 只考虑売上高和労務費以下的部分
            df_analysis_sub = df_analysis_sub.drop(df.index[1:22])
            df_analysis_sub = df_analysis_sub.drop(df.index[46:48])
            df_analysis_sub.reset_index(drop=True, inplace=True)

            df_analysis_sub.set_index("勘定科目", inplace=True)

            row_counter = 1
            num_counter = 1

            # 循环遍历每一个物件，写入excel分析表中
            for i in range(len(analysis_list)):

                project = df_analysis_sub[analysis_list[i]]
                project = project[(project > 300000) | (project < -300000)]

                # 写入数据
                ws2.cell(row=row_counter, column=5, value=num_counter).font = Font(name='等线')
                ws2.cell(row=row_counter, column=6, value=project.name).font = Font(name='等线')
                different = float(df_analysis[df_analysis['物件名'] == project.name]["損益予実差"].iloc[0])

                if different < 0:
                    ws2.cell(row=row_counter, column=7, value=different).font = Font(color="FF0000", name='等线')
                else:
                    ws2.cell(row=row_counter, column=7, value=different).font = Font(name='等线')
                row_counter += 1

                sum_counter = 0

                # 循环写入每一个科目
                for row_idx, row_values in enumerate(project, 0):

                    index = project.index[row_idx]

                    # 判断是否是売上高还是费用，如果是费用就相反数，如果是売上高就原封不动
                    if index == "売上高":
                        pass
                    else:
                        row_values = -row_values

                    sum_counter += row_values
                    ws2.cell(row=row_counter, column=6, value=index).font = Font(name='等线')
                    ws2.cell(row=row_counter, column=7, value=row_values).font = Font(name='等线')
                    if row_values >= 0:
                        ws2.cell(row=row_counter, column=7, value=row_values).font = Font(name='等线')
                        ws2.cell(row=row_counter, column=8, value="有利差異").alignment = Alignment(horizontal='center')
                        ws2.cell(row=row_counter, column=8).font = Font(name='等线')
                    if row_values < 0:
                        cell = ws2.cell(row=row_counter, column=7, value=row_values)
                        cell.font = Font(color="FF0000", name='等线')
                        cell = ws2.cell(row=row_counter, column=8, value="不利差異")
                        cell.font = Font(color="FF0000", name='等线')
                        cell.alignment = Alignment(horizontal='center')
                    else:
                        pass
                    row_counter += 1
                # 设置上格线和合计值

                ws2.cell(row=row_counter, column=6).border = Border(top=Side(border_style='thin'))
                ws2.cell(row=row_counter, column=7, value=sum_counter).border = Border(top=Side(border_style='thin'))
                if sum_counter < 0:
                    ws2.cell(row=row_counter, column=7).font = Font(color="FF0000", name='等线')
                else:
                    ws2.cell(row=row_counter, column=7).font = Font(name='等线')
                ws2.cell(row=row_counter, column=8).border = Border(top=Side(border_style='thin'))

                # 换行并让物件计数加一
                num_counter += 1
                row_counter += 3

            # ==============================================
            """
            设置格式
            """
            # 设置数值格式为千分位加逗号
            number_format = '#,##0'
            # 遍历所有单元格，设置数值格式
            for row in ws2.iter_rows():
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = number_format

            # 调整单元格宽度
            adjust_worksheet_width(ws2, 14, 1.4)

            # 在最上面插入两行空行
            ws2.insert_rows(1, 2)
            save_excel_file = os.path.splitext(txt_file)[0] + '.xlsx'

            # 在处理完成后调用下载函数
            self.download_excel_file(workbook=wb, default_file_name=save_excel_file)
            wb.close()
        except Exception as e:
            show_error_popup(f"An error occurred: {str(e)}")

    def process_accounting(self):
        try:
            year, ok1 = QInputDialog.getInt(self, "年の入力", "年を入力してください:", 2023, 1900, 3000, 1)
            month, ok2 = QInputDialog.getInt(self, "月の入力", "月を入力してください:", 1, 1, 12, 1)

            # 如果用户点击了取消按钮，则返回
            if not ok1 or not ok2:
                return
            # 处理 TXT 文件到 Excel 文件的逻辑
            csv_file = self.file_path

            df = pd.read_csv(csv_file, encoding='utf-8', parse_dates=["date"])
            df['year'] = df.date.dt.year
            df['month'] = df.date.dt.month
            df['day'] = df.date.dt.day

            df = df.loc[df['year'] == year]

            df = df.sort_values(by=['date', 'debit'])
            df = df[df["month"] == month].reset_index(drop=True)

            wb = load_workbook("template/template.xlsx")

            account_temp = wb['出納帳モデル']
            account = wb.copy_worksheet(account_temp)
            sheet_name = "{0}月出納帳".format(month)
            account.title = sheet_name

            account.cell(2, 5, "令和{}年{}月本社小口現金出納帳".format(year - 2018, month))

            for i in range(0, len(df)):
                col = df.loc[i]
                # 月の入力
                month = int(col['month'])
                account.cell(6 + i, 1, month)
                # 日の入力
                day = int(col['day'])
                account.cell(6 + i, 2, value=day)

                account.cell(6 + i, 3, value="京都事業所")  # 部門の入力
                account.cell(6 + i, 4, value=col['account'])  # 鑑定科目の入力
                account.cell(6 + i, 5, value=col['description'])  # 摘要の入力
                account.cell(6 + i, 6, value=col['apart'])
                account.cell(6 + i, 7, value=col['members'])

                if col['tax'] == "有り":
                    account.cell(6 + i, 8, value="✓")
                elif col['tax'] == "無し":
                    account.cell(6 + i, 8, value="×")

                if col['invoice'] == "有り":
                    account.cell(6 + i, 9, value="✓")
                elif col['invoice'] == "無し":
                    account.cell(6 + i, 9, value="×")

                # 金額の入力
                pay = col['amount']
                debit = int(col['debit'])
                if debit == 0:
                    account.cell(6 + i, 10, value=pay)
                elif debit == 1:
                    account.cell(6 + i, 11, value=pay)

                # 差额
                balance = df.iloc[0:i + 1, :].loc[df['debit'] == 0]['amount'].sum() \
                          - df.iloc[0:i + 1, :].loc[df['debit'] == 1]['amount'].sum()
                account.cell(6 + i, 12, value=balance)

            # 合計の入力
            account.cell(6 + len(df) + 1, 10, value=df.loc[df['debit'] == 0]['amount'].sum())
            account.cell(6 + len(df) + 1, 11, value=df.loc[df['debit'] == 1]['amount'].sum())
            account.cell(6 + len(df) + 1, 12,
                         value=df.loc[df['debit'] == 0]['amount'].sum() - df.loc[df['debit'] == 1]['amount'].sum())

            # 伝票の処理
            df = df[df['debit'] == 1]

            receipt_temp = wb['伝票モデル']

            for i in range(1, 32):

                record = df[(df["day"] == i) & (df["month"] == month)]
                if not record.empty:
                    sheet_name = "伝票{0}.{1}".format(month, i)
                    receipt = wb.copy_worksheet(receipt_temp)
                    receipt.title = sheet_name

                    for j in range(0, len(record)):
                        col = record.iloc[j, :]

                        # 金額の入力
                        receipt.cell(8 + j, 1, value=col['amount'])
                        # 部門の入力
                        receipt.cell(8 + j, 2, value="京都事業所")
                        # 鑑定科目の入力
                        receipt.cell(8 + j, 3, value=col['account'])
                        # 摘要の入力
                        receipt.cell(8 + j, 4, value=col['description'])

                    # 日付の入力
                    receipt.cell(2, 5, value="令和{0}年".format(year - 2018))
                    receipt.cell(2, 6, value="{0}月{1}日".format(month, i))
                    # 合計の入力
                    receipt.cell(8 + len(record), 1, value=record['amount'].sum())
                    receipt.cell(8 + len(record), 4, value="合計")
                    receipt.cell(8 + len(record), 6, value=record['amount'].sum())
                    # 貸方の入力
                    receipt.cell(8, 5, value="小口現金")
                    receipt.cell(8, 6, value=record['amount'].sum())
                    receipt.cell(5, 6, "金野")

            # テンプレートをキャンセルする
            wb.remove(wb['伝票モデル'])
            wb.remove(wb['出納帳モデル'])
            file_name = "{0}月京都小口現金出納帳令和{1}年.xlsx".format(month, year - 2018)
            # 在处理完成后调用下载函数
            self.download_excel_file(workbook=wb, default_file_name=file_name)
            wb.close()

        except Exception as e:
            show_error_popup(f"An error occurred: {str(e)}")


def main():
    app = QApplication(sys.argv)
    ex = FileDropApp()
    ex.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
